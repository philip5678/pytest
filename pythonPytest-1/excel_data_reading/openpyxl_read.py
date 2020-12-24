import openpyxl
from openpyxl import Workbook

# https://geek-docs.com/python/python-tutorial/python-openpyxl.html
# read, modify and save
book = openpyxl.load_workbook('./test_data.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['B2']
a3 = sheet.cell(row=3, column=1)

print(a1.value)
print(a2.value)
print(a3.value)

print(book.get_sheet_names())
print(book.sheetnames)

active_sheet = book.active
print(type(active_sheet), active_sheet)

sheet = book.get_sheet_by_name("add_test")
print(sheet, sheet.title)
sheet = book['add_test-1']
print(sheet.title)

sheet['B3'] = 11
sheet.cell(row=2, column=2).value = 22

book.save('write2cell.xlsx')
print('--------------------------------------------------')
# ————————————————————————————————————————————————————————————————————————————————————————————————
# create new sheet
book = Workbook()
sheet = book.active

c = 'address'
rows = (
    ('a', 'b', c),
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

book.save('appending.xlsx')
print('--------------------------------------------------')
# ————————————————————————————————————————————————————————————————————————————————————————————————
from openpyxl import Workbook
import time

book = Workbook()
sheet = book.active

sheet['A1'] = 56
sheet['A2'] = 43

now = time.strftime("%x")
sheet['A3'] = now

book.save("sample.xlsx")

book = openpyxl.load_workbook('sample.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)

print(a1.value)
print(a2.value)
print(a3.value)
print('--------------------------------------------------')
# ————————————————————————————————————————————————————————————————————————————————————————————————
from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

print('--------------------------------------------------')
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()
print('--------------------------------------------------')

for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()
print('--------------------------------------------------')

book.save('iterbyrows.xlsx')

# ————————————————————————————————————————————————————————————————————————————————————————————————
book = openpyxl.load_workbook('iterbyrows.xlsx')

sheet = book.active

cells = sheet['A1': 'c6']

for c1, c2, c3 in cells:
    print("{0:8} {1:8} {2:8}".format(c1.value, c2.value, c3.value))
print()
print('--------------------------------------------------')

cells = sheet['A1': 'B6']

for c1, c2 in cells:
    print("{0:8} {1:8}".format(c1.value, c2.value))
print()
print('--------------------------------------------------')
